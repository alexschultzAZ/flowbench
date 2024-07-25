import requests
from datetime import datetime, timezone
import pytz
import openpyxl

def mst_to_unix_timestamp(date_string, time_string):
    mst = pytz.timezone('MST')
    dt = datetime.strptime(f"{date_string} {time_string}", "%Y-%m-%d %H:%M:%S")
    
    dt = mst.localize(dt)
    utc_datetime = dt.astimezone(pytz.utc) 
    # print(utc_datetime)
    return int(utc_datetime.timestamp())

# Define the Prometheus server URL and the query
prometheus_url = 'http://localhost:9090/api/v1/query_range'
query_cpu_usage_time = 'container_cpu_usage_seconds_total{container=~"vidsplit|modect|facextract|facerec"}'
query_memory = 'container_memory_usage_bytes{container=~"vidsplit|modect|facextract|facerec"}'
query_network_transmit = 'container_network_transmit_bytes_total{pod=~"vidsplit-84df68b58f-f64tz|modect-77f8c9445d-268wn|facextract-5ccd499589-6ngfs|facerec-56b687c486-tnp5w"}'
query_network_receive = 'container_network_receive_bytes_total{pod=~"vidsplit-84df68b58f-f64tz|modect-77f8c9445d-268wn|facextract-5ccd499589-6ngfs|facerec-56b687c486-tnp5w"}'
query_cpu_percentage = 'sum (rate (container_cpu_usage_seconds_total{container=~"vidsplit|modect|facextract|facerec"}[60s])) / sum (machine_cpu_cores) * 100'
queries = [query_cpu_percentage]

# Input your local MST time
start_date = '2024-07-25'  # Example date
start_time = '13:54:30'    # Example start time
end_date = '2024-07-25'    # Example date
end_time = '13:55:40'      # Example end time

# Convert MST time to Unix timestamp
start = mst_to_unix_timestamp(start_date, start_time)
end = mst_to_unix_timestamp(end_date, end_time)
step = '10s'

def run_cpu_percentage_query(query, start, end, step):
    response = requests.get(prometheus_url, params={'query': query, 'start': start, 'end': end, 'step': step})
    print(start)
    print(end)
    # Parse the JSON response
    data = response.json()
    print(data)

def run_queries(queries, start, end, step):
    for query in queries:
            
        # Make the request
        response = requests.get(prometheus_url, params={'query': query, 'start': start, 'end': end, 'step': step})
        print(start)
        print(end)
        # Parse the JSON response
        data = response.json()
        print(data)
        wb = openpyxl.Workbook()
        ws_main = wb.active
        ws_main.title = "Summary"

        query_type = ''
        # Write the header for the main sheet
        ws_main.append(["Metric", "Pod Name", "Timestamp", "Value"])

        # Write the data and create subsheets
        for result in data['data']['result']:
            if '__name__' in result['metric']:

                metric = result['metric']['__name__']
            else:
                metric = 'cpu_percentage'
                    
            query_type = metric
            pod_name = result['metric'].get('pod', 'unknown_pod')
            sheet_name = f"{pod_name[:27]}"  # Ensure the sheet name length does not exceed 31 characters
            if sheet_name not in wb.sheetnames:
                ws_sub = wb.create_sheet(title=sheet_name)
                ws_sub.append(["Timestamp", "Value"])
            else:
                ws_sub = wb[sheet_name]
            
            for value in result['values']:
                timestamp = datetime.fromtimestamp(value[0], tz=timezone.utc).astimezone(pytz.timezone('MST'))
                timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
                
                # Write to the main sheet
                ws_main.append([metric, pod_name, timestamp_str, value[1]])
                
                # Write to the subsheet
                ws_sub.append([timestamp_str, value[1]])


        # Save the workbook
        file_name = f'prometheus_data_{query_type}.xlsx'
        wb.save(filename=file_name)

        print(f"Data written to {file_name}")


run_queries(queries=queries, start=start, end=end, step=step)
